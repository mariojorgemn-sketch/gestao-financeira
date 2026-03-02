# app.py — Sistema de Gestão Financeira
# Streamlit Cloud + SQLite + Claude API

import streamlit as st
import sqlite3
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime
from calendar import monthrange
import base64
import json
import re
import io
import pdfplumber

# ══════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Gestão Financeira",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded",
)

# CSS personalizado
st.markdown("""
<style>
    .metric-card {
        background: white;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        text-align: center;
    }
    .metric-label { color: #757575; font-size: 14px; margin-bottom: 4px; }
    .metric-value { font-size: 28px; font-weight: bold; }
    .lucro  { color: #2E7D32; }
    .prejuizo { color: #C62828; }
    .receita-val { color: #2E7D32; }
    .despesa-val { color: #C62828; }
    div[data-testid="stSidebarNav"] { display: none; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# CATEGORIAS
# ══════════════════════════════════════════════════════════════

CATEGORIAS_RECEITA = ["Vendas", "Serviços", "Investimentos", "Comissões", "Outros"]
CATEGORIAS_DESPESA = ["Salários", "Aluguel", "Marketing", "Fornecedores",
                      "Utilidades", "Impostos", "Manutenção", "Outros"]

# ══════════════════════════════════════════════════════════════
# BASE DE DADOS
# ══════════════════════════════════════════════════════════════

DB = "financeiro.db"

def con():
    return sqlite3.connect(DB, check_same_thread=False)

def init_db():
    with con() as c:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS receitas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao TEXT NOT NULL,
                categoria TEXT NOT NULL,
                valor REAL NOT NULL,
                data TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS despesas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao TEXT NOT NULL,
                categoria TEXT NOT NULL,
                valor REAL NOT NULL,
                data TEXT NOT NULL
            );
        """)

init_db()

# ── CRUD helpers ──────────────────────────────────────────────

def query(sql, params=()):
    with con() as c:
        return pd.read_sql_query(sql, c, params=params)

def execute(sql, params=()):
    with con() as c:
        c.execute(sql, params)
        c.commit()

def listar(tabela, ini=None, fim=None):
    if ini and fim:
        return query(f"SELECT * FROM {tabela} WHERE data BETWEEN ? AND ? ORDER BY data DESC", (ini, fim))
    return query(f"SELECT * FROM {tabela} ORDER BY data DESC")

def inserir(tabela, descricao, categoria, valor, data):
    execute(f"INSERT INTO {tabela} (descricao,categoria,valor,data) VALUES (?,?,?,?)",
            (descricao, categoria, valor, data))

def atualizar(tabela, id_, descricao, categoria, valor, data):
    execute(f"UPDATE {tabela} SET descricao=?,categoria=?,valor=?,data=? WHERE id=?",
            (descricao, categoria, valor, data, id_))

def deletar(tabela, id_):
    execute(f"DELETE FROM {tabela} WHERE id=?", (id_,))

def totais(ini=None, fim=None):
    r = listar("receitas", ini, fim)["valor"].sum() if not listar("receitas", ini, fim).empty else 0
    d = listar("despesas", ini, fim)["valor"].sum() if not listar("despesas", ini, fim).empty else 0
    return float(r), float(d), float(r - d)

# ══════════════════════════════════════════════════════════════
# FORMATAÇÃO €
# ══════════════════════════════════════════════════════════════

def eur(valor):
    """Formata valor em euros: 1.234,56 €"""
    return f"{valor:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")

# ══════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════

def exportar_excel(ini=None, fim=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    cab = ["ID", "Data", "Descrição", "Categoria", "Valor (€)"]

    def estilizar(ws, cor_hex):
        ws.append(cab)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=cor_hex)
            cell.alignment = Alignment(horizontal="center")

    ws_r = wb.active
    ws_r.title = "Receitas"
    estilizar(ws_r, "2E7D32")
    for _, row in listar("receitas", ini, fim).iterrows():
        ws_r.append([row.id, row.data, row.descricao, row.categoria, row.valor])

    ws_d = wb.create_sheet("Despesas")
    estilizar(ws_d, "C62828")
    for _, row in listar("despesas", ini, fim).iterrows():
        ws_d.append([row.id, row.data, row.descricao, row.categoria, row.valor])

    tr, td, tl = totais(ini, fim)
    ws_s = wb.create_sheet("Resumo")
    ws_s.append(["Indicador", "Valor (€)"])
    ws_s.append(["Total Receitas", tr])
    ws_s.append(["Total Despesas", td])
    ws_s.append(["Lucro Líquido",  tl])

    for ws in [ws_r, ws_d, ws_s]:
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════
# EXTRAÇÃO LOCAL DE PDF — pdfplumber (gratuito, sem API)
# ══════════════════════════════════════════════════════════════

def extrair_texto_pdf(pdf_bytes):
    """Extrai todo o texto de um PDF usando pdfplumber."""
    texto = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pagina in pdf.pages:
            texto += (pagina.extract_text() or "") + "\n"
    return texto


def parsear_data(texto):
    """Tenta encontrar uma data no texto em vários formatos."""
    padroes = [
        r"\b(\d{4}-\d{2}-\d{2})\b",           # 2024-03-15
        r"\b(\d{2}/\d{2}/\d{4})\b",            # 15/03/2024
        r"\b(\d{2}-\d{2}-\d{4})\b",            # 15-03-2024
        r"\b(\d{1,2}\s+\w+\s+\d{4})\b",        # 15 março 2024
    ]
    formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]
    meses_pt = {
        "janeiro":"01","fevereiro":"02","março":"03","abril":"04",
        "maio":"05","junho":"06","julho":"07","agosto":"08",
        "setembro":"09","outubro":"10","novembro":"11","dezembro":"12"
    }
    for padrao in padroes:
        m = re.search(padrao, texto, re.IGNORECASE)
        if m:
            s = m.group(1)
            for fmt in formatos:
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
            # Tentar formato "15 março 2024"
            for pt, num in meses_pt.items():
                if pt in s.lower():
                    s2 = re.sub(pt, num, s.lower())
                    try:
                        return datetime.strptime(s2, "%d %m %Y").strftime("%Y-%m-%d")
                    except ValueError:
                        pass
    return str(date.today())


def parsear_valor(texto):
    """Extrai o valor total da fatura — procura o maior valor monetário."""
    # Padrões comuns em faturas portuguesas
    padroes_total = [
        r"total\s+(?:a\s+pagar|geral|fatura)[\s:]*([0-9]+[.,][0-9]{2})",
        r"valor\s+total[\s:]*([0-9]+[.,][0-9]{2})",
        r"montante\s+total[\s:]*([0-9]+[.,][0-9]{2})",
        r"total[\s:]+([0-9]+[.,][0-9]{2})\s*€",
        r"([0-9]+[.,][0-9]{2})\s*€",
    ]
    for padrao in padroes_total:
        m = re.search(padrao, texto, re.IGNORECASE)
        if m:
            val = m.group(1).replace(".", "").replace(",", ".")
            try:
                return float(val)
            except ValueError:
                pass
    # Fallback: encontrar todos os valores e devolver o maior
    todos = re.findall(r"\b(\d{1,6}[.,]\d{2})\b", texto)
    valores = []
    for v in todos:
        try:
            valores.append(float(v.replace(".", "").replace(",", ".")))
        except ValueError:
            pass
    return max(valores) if valores else 0.0


def inferir_categoria(texto):
    """Infere a categoria da despesa com base em palavras-chave."""
    texto_lower = texto.lower()
    mapa = {
        "Salários":     ["salário","salario","vencimento","remuneração","rmh","recibo verde"],
        "Aluguel":      ["aluguer","aluguel","arrendamento","renda","alugar"],
        "Marketing":    ["marketing","publicidade","campanha","anúncio","google ads","facebook"],
        "Fornecedores": ["fornecedor","material","mercadoria","produto","stock","compra"],
        "Utilidades":   ["electricidade","eletricidade","água","agua","gás","gas","internet",
                         "telefone","telecomunicações","edp","galp","nos","meo","vodafone"],
        "Impostos":     ["iva","imposto","taxa","irs","irc","segurança social","ss","at -"],
        "Manutenção":   ["manutenção","manutencao","reparação","reparacao","serviço técnico"],
    }
    for categoria, palavras in mapa.items():
        for palavra in palavras:
            if palavra in texto_lower:
                return categoria
    return "Outros"


def extrair_descricao(texto, nome_ficheiro):
    """Tenta extrair uma descrição curta da fatura."""
    # Procurar linhas com "fatura", "referência", "descrição"
    for linha in texto.split("\n"):
        linha = linha.strip()
        if any(k in linha.lower() for k in ["fatura nº","fatura n.","ref.","referência","documento"]):
            if len(linha) > 5:
                return linha[:80]
    # Fallback: usar o nome do ficheiro sem extensão
    return nome_ficheiro.replace(".pdf", "").replace("_", " ")


def analisar_pdf_local(pdf_bytes, nome_ficheiro):
    """
    Extrai dados financeiros de um PDF localmente com pdfplumber.
    Devolve lista de registos no mesmo formato da versão IA.
    """
    texto = extrair_texto_pdf(pdf_bytes)

    if not texto.strip():
        raise ValueError("PDF sem texto extraível (pode ser uma imagem digitalizada).")

    data      = parsear_data(texto)
    valor     = parsear_valor(texto)
    categoria = inferir_categoria(texto)
    descricao = extrair_descricao(texto, nome_ficheiro)

    # Tentar extrair múltiplas linhas de extrato bancário
    # Padrão típico BPI/CGD: "DD/MM/YYYY Descrição -123,45"
    registos = []
    padrao_extrato = re.compile(
        r"(\d{2}[/-]\d{2}[/-]\d{4})\s+(.+?)\s+([-+]?\d{1,6}[.,]\d{2})\s*€?$",
        re.MULTILINE
    )
    for m in padrao_extrato.finditer(texto):
        try:
            d_str = m.group(1).replace("-", "/")
            d_obj = datetime.strptime(d_str, "%d/%m/%Y").strftime("%Y-%m-%d")
            desc  = m.group(2).strip()[:80]
            val   = float(m.group(3).replace(".", "").replace(",", "."))
            if val == 0:
                continue
            registos.append({
                "data":      d_obj,
                "descricao": desc,
                "categoria": inferir_categoria(desc),
                "valor":     abs(val),
                "_tipo":     "receita" if val > 0 else "despesa",
            })
        except ValueError:
            continue

    # Se não encontrou linhas de extrato, devolve registo único
    if not registos:
        registos = [{
            "data":      data,
            "descricao": descricao,
            "categoria": categoria,
            "valor":     valor,
            "_tipo":     "despesa",
        }]

    return registos

# ══════════════════════════════════════════════════════════════
# SIDEBAR — NAVEGAÇÃO
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 💼 Gestão Financeira")
    st.markdown("---")
    pagina = st.radio("", [
        "📊 Dashboard",
        "💰 Receitas",
        "💸 Despesas",
        "📂 Importar PDF",
        "📋 Relatório Mensal",
    ], label_visibility="collapsed")
    st.markdown("---")
    st.caption("Sistema de Gestão Financeira")

# ══════════════════════════════════════════════════════════════
# PÁGINA: DASHBOARD
# ══════════════════════════════════════════════════════════════

if pagina == "📊 Dashboard":
    st.title("📊 Dashboard")

    # Filtro de período
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        ini = st.date_input("De", value=date(date.today().year, 1, 1), key="dash_ini")
    with col2:
        fim = st.date_input("Até", value=date.today(), key="dash_fim")
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        exportar = st.button("⬇️ Exportar Excel", use_container_width=True)

    ini_str = str(ini)
    fim_str = str(fim)

    if exportar:
        buf = exportar_excel(ini_str, fim_str)
        st.download_button(
            label="📥 Clica para descarregar",
            data=buf,
            file_name=f"relatorio_{ini_str}_{fim_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Cards de resumo
    tr, td, tl = totais(ini_str, fim_str)
    cor_lucro = "lucro" if tl >= 0 else "prejuizo"
    emoji_lucro = "📈" if tl >= 0 else "📉"

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">💰 Total de Receitas</div>
            <div class="metric-value receita-val">{eur(tr)}</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">💸 Total de Despesas</div>
            <div class="metric-value despesa-val">{eur(td)}</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">{emoji_lucro} Lucro Líquido</div>
            <div class="metric-value {cor_lucro}">{eur(tl)}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Gráficos
    col_g1, col_g2 = st.columns(2)

    with col_g1:
        fig_bar = go.Figure(data=[
            go.Bar(name="Receitas", x=["Receitas"], y=[tr],
                   marker_color="#2E7D32"),
            go.Bar(name="Despesas", x=["Despesas"], y=[td],
                   marker_color="#C62828"),
            go.Bar(name="Lucro", x=["Lucro"], y=[abs(tl)],
                   marker_color="#1565C0" if tl >= 0 else "#E53935"),
        ])
        fig_bar.update_layout(
            title="Receitas vs Despesas vs Lucro",
            showlegend=False,
            plot_bgcolor="white",
            yaxis_tickprefix="€",
            height=350,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    with col_g2:
        df_desp = listar("despesas", ini_str, fim_str)
        if not df_desp.empty:
            cat_desp = df_desp.groupby("categoria")["valor"].sum().reset_index()
            fig_pie = px.pie(cat_desp, values="valor", names="categoria",
                             title="Despesas por Categoria",
                             color_discrete_sequence=px.colors.qualitative.Set3)
            fig_pie.update_layout(height=350)
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.info("Sem despesas no período selecionado.")

    # Gráfico de evolução mensal
    df_r = listar("receitas", ini_str, fim_str)
    df_d = listar("despesas", ini_str, fim_str)

    if not df_r.empty or not df_d.empty:
        if not df_r.empty:
            df_r["mes"] = df_r["data"].str[:7]
            r_mensal = df_r.groupby("mes")["valor"].sum().reset_index()
            r_mensal.columns = ["mes", "receitas"]
        else:
            r_mensal = pd.DataFrame(columns=["mes", "receitas"])

        if not df_d.empty:
            df_d["mes"] = df_d["data"].str[:7]
            d_mensal = df_d.groupby("mes")["valor"].sum().reset_index()
            d_mensal.columns = ["mes", "despesas"]
        else:
            d_mensal = pd.DataFrame(columns=["mes", "despesas"])

        df_mensal = pd.merge(r_mensal, d_mensal, on="mes", how="outer").fillna(0).sort_values("mes")

        fig_linha = go.Figure()
        if "receitas" in df_mensal.columns:
            fig_linha.add_trace(go.Scatter(x=df_mensal["mes"], y=df_mensal["receitas"],
                                            name="Receitas", line=dict(color="#2E7D32", width=2),
                                            fill="tozeroy", fillcolor="rgba(46,125,50,0.1)"))
        if "despesas" in df_mensal.columns:
            fig_linha.add_trace(go.Scatter(x=df_mensal["mes"], y=df_mensal["despesas"],
                                            name="Despesas", line=dict(color="#C62828", width=2),
                                            fill="tozeroy", fillcolor="rgba(198,40,40,0.1)"))
        fig_linha.update_layout(
            title="Evolução Mensal",
            plot_bgcolor="white",
            yaxis_tickprefix="€",
            height=300,
        )
        st.plotly_chart(fig_linha, use_container_width=True)

# ══════════════════════════════════════════════════════════════
# PÁGINA: RECEITAS / DESPESAS (reutilizável)
# ══════════════════════════════════════════════════════════════

elif pagina in ("💰 Receitas", "💸 Despesas"):
    tipo      = "receitas" if pagina == "💰 Receitas" else "despesas"
    cats      = CATEGORIAS_RECEITA if tipo == "receitas" else CATEGORIAS_DESPESA
    cor_label = "receita-val" if tipo == "receitas" else "despesa-val"
    emoji     = "💰" if tipo == "receitas" else "💸"

    st.title(f"{emoji} {tipo.capitalize()}")

    # ── Formulário de novo registo ──
    with st.expander("➕ Adicionar novo registo", expanded=False):
        with st.form(f"form_{tipo}", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns([2, 3, 2, 2])
            with c1:
                f_data = st.date_input("Data", value=date.today())
            with c2:
                f_desc = st.text_input("Descrição")
            with c3:
                f_cat  = st.selectbox("Categoria", cats)
            with c4:
                f_val  = st.number_input("Valor (€)", min_value=0.0,
                                          step=0.01, format="%.2f")
            submitted = st.form_submit_button("💾 Guardar", use_container_width=True)
            if submitted:
                if not f_desc:
                    st.error("Descrição obrigatória.")
                elif f_val <= 0:
                    st.error("Valor tem de ser maior que zero.")
                else:
                    inserir(tipo, f_desc, f_cat, f_val, str(f_data))
                    st.success("Registo adicionado com sucesso!")
                    st.rerun()

    # ── Filtro de período ──
    st.markdown("#### 🔍 Filtrar por período")
    fc1, fc2, fc3 = st.columns([2, 2, 1])
    with fc1:
        f_ini = st.date_input("De", value=date(date.today().year, 1, 1), key=f"ini_{tipo}")
    with fc2:
        f_fim = st.date_input("Até", value=date.today(), key=f"fim_{tipo}")
    with fc3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Limpar filtro", key=f"limpar_{tipo}"):
            st.rerun()

    df = listar(tipo, str(f_ini), str(f_fim))

    if df.empty:
        st.info("Nenhum registo encontrado.")
    else:
        # Total do período
        total = df["valor"].sum()
        st.markdown(f"**Total no período: <span class='{cor_label}'>{eur(total)}</span>**",
                    unsafe_allow_html=True)

        # Tabela com edição/exclusão
        st.markdown("---")
        for _, row in df.iterrows():
            col_info, col_val, col_ed, col_del = st.columns([4, 2, 1, 1])
            with col_info:
                st.markdown(f"**{row['descricao']}** · {row['categoria']} · {row['data']}")
            with col_val:
                st.markdown(f"<span class='{cor_label}'><b>{eur(row['valor'])}</b></span>",
                            unsafe_allow_html=True)
            with col_ed:
                if st.button("✏️", key=f"ed_{tipo}_{row['id']}",
                             help="Editar"):
                    st.session_state[f"editar_{tipo}"] = row.to_dict()
            with col_del:
                if st.button("🗑️", key=f"del_{tipo}_{row['id']}",
                             help="Eliminar"):
                    deletar(tipo, row["id"])
                    st.rerun()

        # Modal de edição
        chave_ed = f"editar_{tipo}"
        if chave_ed in st.session_state and st.session_state[chave_ed]:
            reg = st.session_state[chave_ed]
            st.markdown("---")
            st.markdown("#### ✏️ Editar registo")
            with st.form(f"edit_form_{tipo}"):
                ec1, ec2, ec3, ec4 = st.columns([2, 3, 2, 2])
                with ec1:
                    e_data = st.date_input("Data",
                                           value=datetime.strptime(reg["data"], "%Y-%m-%d").date())
                with ec2:
                    e_desc = st.text_input("Descrição", value=reg["descricao"])
                with ec3:
                    idx = cats.index(reg["categoria"]) if reg["categoria"] in cats else 0
                    e_cat = st.selectbox("Categoria", cats, index=idx)
                with ec4:
                    e_val = st.number_input("Valor (€)", value=float(reg["valor"]),
                                             min_value=0.0, step=0.01, format="%.2f")
                sc1, sc2 = st.columns(2)
                with sc1:
                    if st.form_submit_button("💾 Guardar alterações", use_container_width=True):
                        atualizar(tipo, reg["id"], e_desc, e_cat, e_val, str(e_data))
                        del st.session_state[chave_ed]
                        st.rerun()
                with sc2:
                    if st.form_submit_button("❌ Cancelar", use_container_width=True):
                        del st.session_state[chave_ed]
                        st.rerun()

# ══════════════════════════════════════════════════════════════
# PÁGINA: IMPORTAR PDF
# ══════════════════════════════════════════════════════════════

elif pagina == "📂 Importar PDF":
    st.title("📂 Importar Faturas PDF")

    # Chave API
    with st.expander("🔑 Configurar chave da API Anthropic", expanded=False):
        api_key_input = st.text_input(
            "Chave API (sk-ant-...)",
            type="password",
            value=st.session_state.get("api_key", ""),
            help="Obtém a chave em console.anthropic.com"
        )
        if st.button("Guardar chave"):
            st.session_state["api_key"] = api_key_input
            st.success("Chave guardada para esta sessão!")

    api_key = st.session_state.get("api_key", "")

    if not api_key:
        st.warning("⚠️ Introduz a tua chave da API Anthropic para usar esta funcionalidade.")
        st.markdown("""
        **Como obter a chave:**
        1. Vai a [console.anthropic.com](https://console.anthropic.com)
        2. Cria uma conta de programador
        3. Vai a **API Keys** → **Create Key**
        4. Copia a chave e cola no campo acima
        """)
    else:
        # Upload de PDFs
        ficheiros = st.file_uploader(
            "Seleciona uma ou mais faturas em PDF",
            type=["pdf"],
            accept_multiple_files=True
        )

        if ficheiros:
            st.markdown(f"**{len(ficheiros)} ficheiro(s) selecionado(s)**")

            if st.button("🤖 Analisar com IA", use_container_width=True, type="primary"):
                todos_registos = []
                barra = st.progress(0, text="A analisar faturas...")

                for i, f in enumerate(ficheiros):
                    barra.progress((i) / len(ficheiros), text=f"A analisar: {f.name}...")
                    try:
                        resultado = analisar_pdf_com_ia(f.read(), f.name, api_key)
                        tipo_doc  = resultado.get("tipo", "despesa")
                        for reg in resultado.get("registos", []):
                            reg["_ficheiro"]  = f.name
                            reg["_tipo"]      = tipo_doc
                            reg["_confirmar"] = True
                            reg.setdefault("data",      str(date.today()))
                            reg.setdefault("descricao", "Sem descrição")
                            reg.setdefault("categoria", "Outros")
                            reg.setdefault("valor",     0.0)
                            todos_registos.append(reg)
                    except Exception as e:
                        st.error(f"Erro em '{f.name}': {e}")

                barra.progress(1.0, text="Análise concluída!")
                st.session_state["registos_pdf"] = todos_registos

            # Mostrar registos extraídos para confirmação
            if "registos_pdf" in st.session_state and st.session_state["registos_pdf"]:
                st.markdown("---")
                st.markdown("### 📋 Registos extraídos — revê e confirma")
                st.caption("Edita os campos se necessário e marca os que queres guardar.")

                regs = st.session_state["registos_pdf"]
                regs_atualizados = []

                for i, reg in enumerate(regs):
                    tipo_reg = reg.get("_tipo", "despesa")
                    cats_reg = CATEGORIAS_RECEITA if tipo_reg == "receitas" else CATEGORIAS_DESPESA

                    with st.container():
                        col_chk, col_fich, col_tipo, col_data, col_desc, col_cat, col_val = st.columns([0.5, 1.5, 1, 1.5, 3, 1.5, 1.5])

                        with col_chk:
                            confirmar = st.checkbox("", value=reg.get("_confirmar", True),
                                                     key=f"chk_{i}")
                        with col_fich:
                            st.caption(reg.get("_ficheiro", ""))
                        with col_tipo:
                            novo_tipo = st.selectbox("", ["despesa", "receita"],
                                                      index=0 if tipo_reg == "despesa" else 1,
                                                      key=f"tipo_{i}", label_visibility="collapsed")
                        with col_data:
                            try:
                                d_val = datetime.strptime(reg.get("data", str(date.today())), "%Y-%m-%d").date()
                            except Exception:
                                d_val = date.today()
                            nova_data = st.date_input("", value=d_val, key=f"data_{i}",
                                                       label_visibility="collapsed")
                        with col_desc:
                            nova_desc = st.text_input("", value=reg.get("descricao", ""),
                                                       key=f"desc_{i}", label_visibility="collapsed")
                        with col_cat:
                            cats_cur = CATEGORIAS_RECEITA if novo_tipo == "receita" else CATEGORIAS_DESPESA
                            cat_val  = reg.get("categoria", cats_cur[0])
                            if cat_val not in cats_cur:
                                cat_val = cats_cur[0]
                            nova_cat = st.selectbox("", cats_cur,
                                                     index=cats_cur.index(cat_val),
                                                     key=f"cat_{i}", label_visibility="collapsed")
                        with col_val:
                            novo_val = st.number_input("", value=float(reg.get("valor", 0)),
                                                        min_value=0.0, step=0.01, format="%.2f",
                                                        key=f"val_{i}", label_visibility="collapsed")

                        regs_atualizados.append({
                            "_ficheiro": reg.get("_ficheiro"),
                            "_tipo":     novo_tipo,
                            "_confirmar": confirmar,
                            "data":      str(nova_data),
                            "descricao": nova_desc,
                            "categoria": nova_cat,
                            "valor":     novo_val,
                        })

                    st.divider()

                st.session_state["registos_pdf"] = regs_atualizados
                confirmados = [r for r in regs_atualizados if r["_confirmar"]]
                st.markdown(f"**{len(confirmados)} de {len(regs_atualizados)} registos selecionados**")

                if st.button("💾 Guardar registos confirmados", type="primary", use_container_width=True):
                    guardados = 0
                    for reg in confirmados:
                        tabela = "receitas" if reg["_tipo"] == "receita" else "despesas"
                        inserir(tabela, reg["descricao"], reg["categoria"],
                                reg["valor"], reg["data"])
                        guardados += 1
                    st.success(f"✅ {guardados} registo(s) guardados com sucesso!")
                    del st.session_state["registos_pdf"]
                    st.rerun()

# ══════════════════════════════════════════════════════════════
# PÁGINA: RELATÓRIO MENSAL
# ══════════════════════════════════════════════════════════════

elif pagina == "📋 Relatório Mensal":
    st.title("📋 Relatório Mensal")

    rc1, rc2 = st.columns(2)
    with rc1:
        mes = st.selectbox("Mês", range(1, 13),
                            format_func=lambda m: ["Janeiro","Fevereiro","Março","Abril",
                                                    "Maio","Junho","Julho","Agosto",
                                                    "Setembro","Outubro","Novembro","Dezembro"][m-1],
                            index=date.today().month - 1)
    with rc2:
        ano = st.number_input("Ano", min_value=2000, max_value=2100,
                               value=date.today().year, step=1)

    ultimo_dia = monthrange(ano, mes)[1]
    ini_rel    = f"{ano}-{mes:02d}-01"
    fim_rel    = f"{ano}-{mes:02d}-{ultimo_dia}"

    tr, td, tl = totais(ini_rel, fim_rel)
    cor_lucro  = "lucro" if tl >= 0 else "prejuizo"
    emoji_l    = "📈" if tl >= 0 else "📉"

    st.markdown("---")

    # Cards
    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">💰 Total Receitas</div>
            <div class="metric-value receita-val">{eur(tr)}</div>
        </div>""", unsafe_allow_html=True)
    with m2:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">💸 Total Despesas</div>
            <div class="metric-value despesa-val">{eur(td)}</div>
        </div>""", unsafe_allow_html=True)
    with m3:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">{emoji_l} Lucro Líquido</div>
            <div class="metric-value {cor_lucro}">{eur(tl)}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Detalhe por categoria
    col_r, col_d = st.columns(2)

    with col_r:
        st.markdown("#### 💰 Receitas por categoria")
        df_r = listar("receitas", ini_rel, fim_rel)
        if not df_r.empty:
            cat_r = df_r.groupby("categoria")["valor"].sum().reset_index().sort_values("valor", ascending=False)
            for _, row in cat_r.iterrows():
                cc1, cc2 = st.columns([3, 2])
                with cc1:
                    st.markdown(f"• {row['categoria']}")
                with cc2:
                    st.markdown(f"<span class='receita-val'><b>{eur(row['valor'])}</b></span>",
                                unsafe_allow_html=True)
        else:
            st.info("Sem receitas neste mês.")

    with col_d:
        st.markdown("#### 💸 Despesas por categoria")
        df_d = listar("despesas", ini_rel, fim_rel)
        if not df_d.empty:
            cat_d = df_d.groupby("categoria")["valor"].sum().reset_index().sort_values("valor", ascending=False)
            for _, row in cat_d.iterrows():
                dc1, dc2 = st.columns([3, 2])
                with dc1:
                    st.markdown(f"• {row['categoria']}")
                with dc2:
                    st.markdown(f"<span class='despesa-val'><b>{eur(row['valor'])}</b></span>",
                                unsafe_allow_html=True)
        else:
            st.info("Sem despesas neste mês.")

    st.markdown("---")

    # Exportar Excel do mês
    buf = exportar_excel(ini_rel, fim_rel)
    meses_pt = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    st.download_button(
        label=f"⬇️ Exportar {meses_pt[mes-1]} {ano} para Excel",
        data=buf,
        file_name=f"relatorio_{meses_pt[mes-1]}_{ano}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
